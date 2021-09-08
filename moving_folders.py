import os
import shutil 
import openpyxl
import glob

wb = openpyxl.load_workbook(filename='urls_2.xlsx')
ws = wb['Лист1']

# origins = list(set([col.value for col in list(ws['B'])[:3]]))

# for origin in origins:
#     if origin is not None:

for origin, inn in zip(list(ws['B'])[3:], list(ws['C'])[3:]):
    if origin is not None:
        if not os.path.exists(origin):
            os.mkdir(origin)
        print(origin.value, inn.value)
        sourcedirs = glob.glob('/output/{}_*'.format(inn))
        for sourcedir in sourcedirs:
            distdir = os.path.join(origin, os.path.dirname(sourcedir))
            shutil.move(sourcedir, distdir)
